"""Read-only, validated readers for the three master workbooks."""
from __future__ import annotations
import re
from pathlib import Path
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load(path: Path) -> openpyxl.Workbook:
    if not path.exists():
        raise FileNotFoundError(f"Master file not found: {path}")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _norm_col(name: str) -> str:
    """Strip parentheticals and normalise to lower_snake."""
    name = re.sub(r"\s*\([^)]*\)", "", name)
    return name.strip().lower().replace(" ", "_").replace("/", "_").replace("-", "_")


def _sheet(wb: openpyxl.Workbook, name: str, path: Path) -> openpyxl.worksheet.worksheet.Worksheet:
    if name not in wb.sheetnames:
        raise ValueError(f"Required sheet '{name}' missing from {path}")
    return wb[name]


def _rows_as_dicts(ws, required: list[str], path: Path, sheet: str) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet '{sheet}' in {path} is empty")
    raw_headers = [str(h) if h is not None else "" for h in rows[0]]
    headers = [_norm_col(h) for h in raw_headers]
    for col in required:
        norm = _norm_col(col)
        if norm not in headers:
            raise ValueError(
                f"Required column '{col}' missing from sheet '{sheet}' in {path}. "
                f"Found: {raw_headers}"
            )
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def _str(v: Any, default: str = "") -> str:
    return str(v).strip() if v is not None else default


def _bool_yes(v: Any) -> bool:
    return _str(v).lower() in ("yes", "y", "true", "1")


def _float(v: Any, default: float = 0.0) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _int(v: Any, default: int = 0) -> int:
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


# ---------------------------------------------------------------------------
# Source registry
# ---------------------------------------------------------------------------

SOURCE_REQUIRED_COLS = [
    "Source Name", "Type", "Active", "URL",
    "Trust Score (1-5)", "Signal Score (1-5)", "Noise Score (1-5)",
    "Priority Tier (1-5)", "Include in Morning Digest", "Include in Afternoon Digest",
]


def read_sources(path: Path) -> list[dict]:
    """Return list of raw source dicts from the registry. Active filtering is the caller's job."""
    wb = _load(path)
    ws = _sheet(wb, "Sources", path)
    rows = _rows_as_dicts(ws, SOURCE_REQUIRED_COLS, path, "Sources")
    wb.close()
    sources = []
    for r in rows:
        sources.append({
            "name": _str(r.get("source_name")),
            "type": _str(r.get("type")),
            "sector": _str(r.get("sector")),
            "adjacent_domain": _str(r.get("adjacent_domain")),
            "active": _bool_yes(r.get("active")),
            "url": _str(r.get("url")),
            "linkedin_url": _str(r.get("linkedin_url")),
            "asx_ticker": _str(r.get("asx_ticker")),
            "primary_or_secondary": _str(r.get("primary_or_secondary")),
            "official_status": _str(r.get("official_status")),
            "trust_score": _float(r.get("trust_score"), 3.0),
            "signal_score": _float(r.get("signal_score"), 3.0),
            "noise_score": _float(r.get("noise_score"), 3.0),
            "priority_tier": _int(r.get("priority_tier"), 3),
            "include_morning": _bool_yes(r.get("include_in_morning_digest", "yes")),
            "include_afternoon": _bool_yes(r.get("include_in_afternoon_digest", "yes")),
            "notes": _str(r.get("notes")),
            "country": _str(r.get("country")),
            "hq_state": _str(r.get("hq_state_territory")),
        })
    return sources


# ---------------------------------------------------------------------------
# kestrel_config.xlsx
# ---------------------------------------------------------------------------

def read_kestrel_config(path: Path) -> dict:
    """Return dict with keys: kpmg_tags, domain_tags, quotes, filters, escalation,
    writing_style_rules, audience."""
    wb = _load(path)
    result = {}

    # Categories_KPMG
    ws = _sheet(wb, "Categories_KPMG", path)
    rows = _rows_as_dicts(ws, ["Tag", "Description", "Active"], path, "Categories_KPMG")
    result["kestrel_tags"] = [
        {"tag": _str(r["tag"]), "description": _str(r["description"])}
        for r in rows if _bool_yes(r.get("active"))
    ]

    # Categories_Domain
    ws = _sheet(wb, "Categories_Domain", path)
    rows = _rows_as_dicts(ws, ["Tag", "Description", "Active"], path, "Categories_Domain")
    result["domain_tags"] = [
        {"tag": _str(r["tag"]), "description": _str(r["description"])}
        for r in rows if _bool_yes(r.get("active"))
    ]

    # Quotes
    ws = _sheet(wb, "Quotes", path)
    rows = _rows_as_dicts(ws, ["Quote", "Author", "Active"], path, "Quotes")
    result["quotes"] = [
        {
            "quote": _str(r["quote"]),
            "author": _str(r["author"]),
            "theme": _str(r.get("theme", "general")) or "general",
        }
        for r in rows if _bool_yes(r.get("active"))
    ]

    # Filters  (Key / Value / Notes)
    ws = _sheet(wb, "Filters", path)
    rows = _rows_as_dicts(ws, ["Key", "Value"], path, "Filters")
    result["filters"] = {_str(r["key"]): _str(r["value"]) for r in rows if r.get("key")}

    # Escalation
    ws = _sheet(wb, "Escalation", path)
    rows = _rows_as_dicts(ws, ["Trigger", "Keywords", "Action", "Active"], path, "Escalation")
    result["escalation"] = [
        {
            "trigger": _str(r["trigger"]),
            "keywords": [k.strip() for k in _str(r["keywords"]).split(",") if k.strip()],
            "action": _str(r["action"]),
        }
        for r in rows if _bool_yes(r.get("active"))
    ]

    # Writing_Style
    ws = _sheet(wb, "Writing_Style", path)
    rows = _rows_as_dicts(ws, ["Rule_Group", "Rule", "Active"], path, "Writing_Style")
    result["writing_style_rules"] = [
        {"group": _str(r["rule_group"]), "rule": _str(r["rule"])}
        for r in rows if _bool_yes(r.get("active"))
    ]

    # Audience
    ws = _sheet(wb, "Audience", path)
    rows = _rows_as_dicts(ws, ["Key", "Value"], path, "Audience")
    result["audience"] = {_str(r["key"]): _str(r["value"]) for r in rows if r.get("key")}

    wb.close()
    return result


# ---------------------------------------------------------------------------
# subscribers.xlsx
# ---------------------------------------------------------------------------

def read_subscribers(path: Path) -> list[dict]:
    if not path.exists():
        return []
    wb = _load(path)
    ws = wb.active
    sheet_name = ws.title
    rows = _rows_as_dicts(ws, ["Name", "Email", "Subscription Preference"], path, sheet_name)
    wb.close()
    return [
        {
            "name": _str(r.get("name")),
            "email": _str(r.get("email")),
            "preference": _str(r.get("subscription_preference")),
        }
        for r in rows if _str(r.get("email"))
    ]
