"""Kestrel website server — FastAPI app serving the landing page and
handling subscribe / unsubscribe API routes.

Reads and writes data/subscribers.xlsx (relative to project root, one level
above this file).  Run from the project root:

    .venv/Scripts/python.exe -m uvicorn website.server:app --host 0.0.0.0 --port 8080

Or from the website/ directory:

    python server.py
"""
from __future__ import annotations

import logging
import re
from pathlib import Path

import openpyxl
import uvicorn
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

log = logging.getLogger(__name__)

# ── Paths ──────────────────────────────────────────────────────────────────
HERE         = Path(__file__).parent                   # website/
PROJECT_ROOT = HERE.parent                             # kestrel/
DATA_DIR     = PROJECT_ROOT / "data"
SUBS_PATH    = DATA_DIR / "subscribers.xlsx"

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# ── App ────────────────────────────────────────────────────────────────────
app = FastAPI(title="Kestrel", docs_url=None, redoc_url=None)
app.mount("/static", StaticFiles(directory=HERE / "static"), name="static")


# ── Helpers ────────────────────────────────────────────────────────────────
def _ensure_workbook() -> openpyxl.Workbook:
    """Return the subscribers workbook, creating it with headers if missing."""
    if SUBS_PATH.exists():
        return openpyxl.load_workbook(SUBS_PATH)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Email", "Subscription Preference"])
    wb.save(SUBS_PATH)
    return wb


def _find_row(ws, email: str):
    """Return the row tuple (row_idx, name, email, pref) or None."""
    email_lower = email.lower()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[1] and str(row[1]).strip().lower() == email_lower:
            return i, row
    return None, None


# ── Models ─────────────────────────────────────────────────────────────────
class SubscribeRequest(BaseModel):
    name: str = ""
    email: str


class UnsubscribeRequest(BaseModel):
    email: str


# ── Routes ─────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def index():
    return FileResponse(HERE / "index.html")


@app.get("/unsubscribe", response_class=HTMLResponse)
async def unsubscribe_page():
    return FileResponse(HERE / "unsubscribe.html")


@app.post("/api/subscribe")
async def subscribe(req: SubscribeRequest):
    email = req.email.strip()
    if not email or not _EMAIL_RE.match(email):
        raise HTTPException(status_code=422, detail="Please provide a valid email address.")

    wb = _ensure_workbook()
    ws = wb.active

    row_idx, existing = _find_row(ws, email)
    if row_idx:
        pref = (existing[2] or "").lower()
        if pref != "unsubscribed":
            return JSONResponse({"status": "already_subscribed",
                                 "message": "You're already subscribed — look for Kestrel in your inbox."})
        # Re-subscribe
        ws.cell(row=row_idx, column=3).value = "active"
        wb.save(SUBS_PATH)
        log.info("Re-subscribed: %s", email)
        return JSONResponse({"status": "ok",
                             "message": "Welcome back — you've been re-subscribed to the daily brief."})

    ws.append([req.name.strip(), email, "active"])
    wb.save(SUBS_PATH)
    log.info("New subscriber: %s", email)
    return JSONResponse({"status": "ok",
                         "message": "You're subscribed. Kestrel will arrive in your inbox tomorrow morning."})


@app.post("/api/unsubscribe")
async def unsubscribe(req: UnsubscribeRequest):
    email = req.email.strip()
    if not email or not _EMAIL_RE.match(email):
        raise HTTPException(status_code=422, detail="Please provide a valid email address.")

    wb = _ensure_workbook()
    ws = wb.active

    row_idx, existing = _find_row(ws, email)
    if not row_idx:
        return JSONResponse(status_code=404,
                            content={"status": "not_found",
                                     "message": "That email isn't on our list. Nothing to unsubscribe."})

    ws.cell(row=row_idx, column=3).value = "unsubscribed"
    wb.save(SUBS_PATH)
    log.info("Unsubscribed: %s", email)
    return JSONResponse({"status": "ok",
                         "message": "Done — you've been unsubscribed from the daily brief."})


# ── Dev runner ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    uvicorn.run("server:app", host="127.0.0.1", port=8080, reload=True)
