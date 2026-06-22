"""One-time source audit: discover the working collection method for every source.

Runs all active sources in parallel, records the first method that yields items,
and writes 'feed: <url>' / 'gnews: <query>' / 'method: scrape' hints back to
australian_defence_source_universe.xlsx so subsequent runs skip discovery entirely.

Usage:
    python -m kestrel audit [--workers 20] [--timeout 20] [--hours 48] [--dry-run]
    python scripts/audit_sources.py
"""
from __future__ import annotations
import concurrent.futures
import logging
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

# Allow running as a plain script
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import openpyxl

from kestrel.config import load_config
from kestrel.models import Source, Window
from kestrel.collectors.protocol import parse_notes_hint
from kestrel.collectors.rss import RSSCollector, _discover_feed
from kestrel.collectors.scrape import HTMLScrapeCollector
from kestrel.collectors.google_news import GoogleNewsCollector

logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(name)s - %(message)s")
log = logging.getLogger("audit")


def _audit_one(source: Source, window: Window, timeout: int) -> dict:
    """Test a source with all methods. Returns audit result dict."""
    result = {
        "name": source.name,
        "type": source.type,
        "url": source.url,
        "method": "none",
        "feed_url": "",
        "item_count": 0,
        "duration_s": 0.0,
        "notes_update": "",
    }

    if source.type.lower() == "linkedin":
        result["method"] = "skip_linkedin"
        return result

    if source.type.upper() in ("ASX", "AUSTENDER"):
        result["method"] = source.type.lower()
        return result

    # --- Try RSS ---
    t0 = time.time()
    try:
        rss = RSSCollector(timeout=timeout)
        items = rss.collect(source, window)
        if items:
            feed_url = items[0].raw_meta.get("feed_url", "")
            result.update(method="rss", feed_url=feed_url,
                          item_count=len(items), duration_s=round(time.time() - t0, 1))
            if feed_url and f"feed:{feed_url}" not in source.notes and feed_url not in source.notes:
                result["notes_update"] = f"feed:{feed_url}"
            return result
    except Exception:
        pass

    # --- Try HTML scrape ---
    t0 = time.time()
    try:
        items = HTMLScrapeCollector(timeout=timeout).collect(source, window)
        if items:
            result.update(method="scrape", item_count=len(items),
                          duration_s=round(time.time() - t0, 1))
            if "method:" not in source.notes:
                result["notes_update"] = "method:scrape"
            return result
    except Exception:
        pass

    # --- Try Google News ---
    t0 = time.time()
    try:
        gnews = GoogleNewsCollector(timeout=timeout)
        items = gnews.collect(source, window)
        if items:
            from urllib.parse import urlparse
            domain = urlparse(source.url).netloc.lstrip("www.")
            query = parse_notes_hint(source.notes, "gnews") or f"site:{domain}"
            result.update(method="gnews", item_count=len(items),
                          duration_s=round(time.time() - t0, 1))
            if "gnews:" not in source.notes:
                result["notes_update"] = f"gnews:{query}"
            return result
    except Exception:
        pass

    return result


def run_audit(
    project_root: Path,
    workers: int = 20,
    timeout: int = 20,
    lookback_hours: int = 48,
    dry_run: bool = False,
) -> list[dict]:
    cfg = load_config(project_root)
    window = Window(
        start=datetime.now(tz=timezone.utc) - timedelta(hours=lookback_hours),
        end=datetime.now(tz=timezone.utc),
    )

    active_sources = [s for s in cfg.sources if s.active]
    print(f"\nAuditing {len(active_sources)} active sources ({workers} workers, "
          f"{timeout}s timeout, {lookback_hours}h window)...\n")

    results: list[dict] = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(_audit_one, s, window, timeout): s for s in active_sources}
        done = 0
        for future in concurrent.futures.as_completed(futures):
            done += 1
            try:
                r = future.result()
                results.append(r)
                icon = {"rss": "[rss]  ", "scrape": "[web]  ", "gnews": "[gnews]",
                        "none": "[none] ", "skip_linkedin": "[skip] "}.get(r["method"], "[?]    ")
                print(f"  [{done:3d}/{len(active_sources)}] {icon} {r['method']:12s} "
                      f"{r['item_count']:3d}items  {r['duration_s']:5.1f}s  {r['name'][:50]}")
            except Exception as exc:
                src = futures[future]
                print(f"  [{done:3d}/{len(active_sources)}] [ERR]  ERROR        "
                      f"  0items    0.0s  {src.name}: {exc}")

    # Summary
    from collections import Counter
    counts = Counter(r["method"] for r in results)
    updates = [r for r in results if r["notes_update"]]
    print(f"\n{'='*60}")
    print(f"Method breakdown: {dict(counts)}")
    print(f"Sources with proposed Notes updates: {len(updates)}")

    if dry_run:
        print("\n-- DRY RUN: proposed Notes updates --")
        for r in updates:
            print(f"  {r['name'][:50]:<50}  +{r['notes_update']}")
        return results

    # Write updates to registry
    if updates:
        xlsx_path = cfg.paths.data_dir / "australian_defence_source_universe.xlsx"
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Sources"]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        notes_col = headers.index("Notes") + 1
        name_col = 1

        name_to_row: dict[str, int] = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name_to_row[str(row[name_col - 1] or "")] = i

        written = 0
        for r in updates:
            row_num = name_to_row.get(r["name"])
            if row_num is None:
                continue
            cell = ws.cell(row=row_num, column=notes_col)
            existing = str(cell.value or "").strip()
            if r["notes_update"] not in existing:
                separator = ", " if existing else ""
                cell.value = existing + separator + r["notes_update"]
                written += 1

        wb.save(xlsx_path)
        print(f"\nWrote {written} Notes updates to {xlsx_path.name}")

    return results


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Audit Kestrel source collection methods")
    p.add_argument("--workers", type=int, default=20)
    p.add_argument("--timeout", type=int, default=20)
    p.add_argument("--hours", type=int, default=48,
                   help="Lookback window in hours (default 48 for broader coverage)")
    p.add_argument("--dry-run", action="store_true",
                   help="Print proposed updates without writing to registry")
    p.add_argument("--project-root", default=None)
    args = p.parse_args()

    root = Path(args.project_root) if args.project_root else Path(__file__).parent.parent
    run_audit(root, args.workers, args.timeout, args.hours, args.dry_run)
