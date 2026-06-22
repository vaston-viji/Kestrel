"""
Apply confirmed feed: and selector: hints to the source registry.
Run from project root:
    .venv/Scripts/python.exe scripts/apply_feed_hints.py [--dry-run]
"""
from __future__ import annotations
import argparse
import re
import sys
from pathlib import Path

import openpyxl

# ---- CONFIRMED HINTS -------------------------------------------------------
# format: "Source Name as in registry" -> "notes value to SET"
# Use "feed: <url>" for RSS/Atom feeds.
# Use "selector: <css>" for HTML scraping.
# Use None to CLEAR an existing broken hint.

HINTS: dict[str, str | None] = {
    # === New RSS feeds (probe-confirmed) ===
    "ASC Pty Ltd":
        "feed: https://www.asc.com.au/feed",

    "ANSTO":
        "feed: https://www.ansto.gov.au/rss.xml",

    "Australian Industry & Defence Network (AIDN)":
        "feed: https://aidn.org.au/feed",

    "Defence SA":
        "feed: https://defencesa.com/feed",

    # === Selector hints for HTML-scrape-only sources ===
    # Australian Defence Magazine — article cards use div.landing-card-title h3 > a
    "Australian Defence Magazine":
        "selector: div.landing-card-title h3 a",

    # Defence Connect — homepage Joomla site; article anchors carry class b-latest-news*
    "Defence Connect":
        "selector: a[class*='b-latest-news']",

    # === Fix broken feed URLs written by earlier auto-discovery ===
    # CSIRO site uses Sitecore CMS; all news is loaded via JS — no working RSS URL found
    "CSIRO": None,  # clear the broken feed: https://www.csiro.au/en/news/rss

    # USSC feed URL returns 404 from all tested paths
    "United States Studies Centre": None,  # clear the broken feed: https://www.ussc.edu.au/feed
}


def _norm_col(name: str) -> str:
    name = re.sub(r"\s*\([^)]*\)", "", name)
    return name.strip().lower().replace(" ", "_").replace("/", "_").replace("-", "_")


def apply_hints(xlsx_path: Path, dry_run: bool) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Sources"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hi: dict[str, int] = {}  # 1-based
    for i, h in enumerate(headers, start=1):
        if h:
            hi[_norm_col(str(h))] = i

    name_col = hi.get("source_name")
    notes_col = hi.get("notes")
    if not name_col or not notes_col:
        print(f"ERROR: can't find name/notes columns. Got: {list(hi)}")
        sys.exit(1)

    matched: set[str] = set()
    changed = 0

    for row in ws.iter_rows(min_row=2):
        name = str(row[name_col - 1].value or "").strip()
        if name not in HINTS:
            continue
        matched.add(name)

        notes_cell = row[notes_col - 1]
        existing = str(notes_cell.value or "").strip()
        desired = HINTS[name]  # str or None

        if desired is None:
            # Clear any broken feed/selector hints
            if existing:
                print(f"  CLEAR  {name}")
                print(f"    was: {existing}")
                if not dry_run:
                    notes_cell.value = None
                changed += 1
            else:
                print(f"  SKIP   {name} (already empty)")
        else:
            # Set / replace the hint
            # Strip any existing feed:/selector: hints to avoid duplicates
            cleaned = re.sub(r"(?i)\s*(?:feed|selector):[^;]+;?\s*", "", existing).strip()
            cleaned = cleaned.strip(";").strip()
            new_val = f"{desired}; {cleaned}" if cleaned else desired

            if notes_cell.value == new_val:
                print(f"  SKIP   {name} (unchanged)")
            else:
                print(f"  UPDATE {name}")
                print(f"    was: {existing!r}")
                print(f"    now: {new_val!r}")
                if not dry_run:
                    notes_cell.value = new_val
                changed += 1

    # Warn about any HINTS entries that didn't match any row
    unmatched = set(HINTS) - matched
    for name in sorted(unmatched):
        print(f"  WARN   '{name}' not found in Sources sheet")

    if not dry_run:
        wb.save(xlsx_path)
        print(f"\nSaved {xlsx_path}  ({changed} rows changed)")
    else:
        print(f"\nDRY RUN — {changed} rows would change")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--xlsx", default="data/australian_defence_source_universe.xlsx")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent.parent
    xlsx = root / args.xlsx
    if not xlsx.exists():
        print(f"ERROR: {xlsx} not found")
        sys.exit(1)

    print(f"Applying hints to {xlsx}")
    apply_hints(xlsx, args.dry_run)
