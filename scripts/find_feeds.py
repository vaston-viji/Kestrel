"""
Probe all active non-LinkedIn sources for RSS/Atom feeds and write
  feed: <url>
hints into the Notes column of data/australian_defence_source_universe.xlsx.

Run from project root:
    .venv/Scripts/python.exe scripts/find_feeds.py [--dry-run]
"""
from __future__ import annotations
import asyncio
import argparse
import logging
import re
import sys
from pathlib import Path
from urllib.parse import urljoin, urlparse

import httpx
import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)-8s %(message)s",
                    stream=sys.stdout)
log = logging.getLogger("find_feeds")

UA = "Kestrel/1.0 (RSS discovery; contact viji.john@quantrim.com)"
HEADERS = {"User-Agent": UA, "Accept": "application/rss+xml, application/atom+xml, text/html"}

# Paths to probe (in order — first hit wins)
FEED_PATHS = [
    "/feed", "/feed/", "/rss.xml", "/atom.xml", "/feed.xml",
    "/news.xml", "/news/feed", "/news/rss.xml", "/news/rss",
    "/blog/feed", "/media-releases.xml", "/publications/rss.xml",
    "/rss", "/rss/", "/feed/rss.xml", "/news-events/feed",
    "/news-events/releases/feed", "/news-events/news/feed",
    "/about/news/feed", "/media/feed", "/updates/feed",
]

# Hardcoded known-good feeds (verified or high-confidence from training data)
# Format: "Source Name from registry" -> "feed_url"
KNOWN_FEEDS: dict[str, str] = {
    # Think tanks (verified)
    "The Strategist, ASPI":                 "https://www.aspistrategist.org.au/feed/",
    "Australian Strategic Policy Institute": "https://www.aspi.org.au/feed",
    "Lowy Institute":                        "https://www.lowyinstitute.org/the-interpreter/rss.xml",
    "United States Studies Centre":          "https://www.ussc.edu.au/feed",

    # Media (verified/high-confidence)
    "InnovationAus":                         "https://www.innovationaus.com/feed",
    "Information Age":                       "https://ia.acs.org.au/feed",
    "Cyber Daily":                           "https://www.cyberdaily.com.au/feed",
    "Australian Aviation":                   "https://australianaviation.com.au/feed",
    "Asia Pacific Defence Reporter":         "https://www.apdronline.com/feed",
    "Australia Defence Association":         "https://australiandefence.org.au/feed",

    # Universities / research
    "UNSW Canberra":                         "https://www.unsw.adfa.edu.au/feed",
    "ANU Defence Institute":                 "https://adi.anu.edu.au/feed",
    "Strategic and Defence Studies Centre, ANU": "https://sdsc.bellschool.anu.edu.au/feed",
    "CSIRO":                                 "https://www.csiro.au/en/news/rss",

    # Government — Australian Government Drupal/GovCMS sites commonly expose /feed
    # These are best-effort; probe will verify/override
    "Defence Media Releases":               "https://www.defence.gov.au/news-events/releases/feed",
    "Defence News":                         "https://www.defence.gov.au/news-events/news/feed",
    "Department of Defence, Home":          "https://www.defence.gov.au/feed",
    "Defence Science and Technology Group": "https://www.dst.defence.gov.au/feed",
    "Australian Submarine Agency, News":    "https://www.asa.gov.au/news/all-news/feed",
    "DFAT Media Releases":                  "https://www.foreignminister.gov.au/minister/penny-wong/news/rss.xml",
    "Cyber.gov.au":                         "https://www.cyber.gov.au/about/news/feed",
    "Critical Infrastructure Security Centre": "https://www.cisc.gov.au/feed",
    "Department of Industry, Science and Resources": "https://www.industry.gov.au/feed",

    # Parliament
    "Parliament, Joint Standing Committee on Foreign Affairs, Defence and Trade":
        "https://www.aph.gov.au/Parliamentary_Business/Committees/Joint/Foreign_Affairs_Defence_and_Trade/RSS",
    "Parliament, Parliamentary Joint Committee on Intelligence and Security":
        "https://www.aph.gov.au/Parliamentary_Business/Committees/Joint/Intelligence_and_Security/RSS",
}

# Sources that are known NOT to have public RSS feeds (skip probing)
NO_FEED_SOURCES = {
    "ASX Announcements platform",   # platform meta-row
    "AusTender",                    # tender DB, no RSS
    "Australian Financial Review",  # paywalled
    "The Australian, Defence and National Security",  # paywalled
    # LinkedIn rows (type=LinkedIn, handled separately by router)
}


async def probe_feed(client: httpx.AsyncClient, base_url: str, timeout: int = 8) -> str | None:
    """Try to find an RSS/Atom feed for a source URL. Returns feed URL or None."""
    root = urlparse(base_url)
    root_url = f"{root.scheme}://{root.netloc}"

    # 1) GET the source URL and look for <link rel="alternate">
    try:
        r = await client.get(base_url, headers=HEADERS, timeout=timeout, follow_redirects=True)
        if r.status_code == 200 and "html" in r.headers.get("content-type", ""):
            found = _extract_feed_link(r.text, base_url)
            if found:
                return found
    except Exception:
        pass

    # 2) Probe common feed paths
    for path in FEED_PATHS:
        for base in (base_url.rstrip("/"), root_url):
            url = urljoin(base + "/", path.lstrip("/"))
            try:
                r = await client.head(url, headers=HEADERS, timeout=timeout, follow_redirects=True)
                ct = r.headers.get("content-type", "").lower()
                if r.status_code == 200 and ("xml" in ct or "rss" in ct or "atom" in ct):
                    return url
                # Some servers return text/html for feeds — do a GET to check
                if r.status_code == 200:
                    r2 = await client.get(url, headers=HEADERS, timeout=timeout, follow_redirects=True)
                    if r2.status_code == 200 and _looks_like_feed(r2.text):
                        return url
            except Exception:
                pass

    return None


def _extract_feed_link(html: str, base_url: str) -> str | None:
    """Find <link rel="alternate" type="application/.../xml"> in HTML."""
    pat = re.compile(
        r'<link[^>]+rel=["\']alternate["\'][^>]+'
        r'type=["\']application/(?:rss|atom)\+xml["\'][^>]*>',
        re.IGNORECASE,
    )
    href_pat = re.compile(r'href=["\']([^"\']+)["\']', re.IGNORECASE)
    for m in pat.finditer(html):
        tag = m.group(0)
        href = href_pat.search(tag)
        if href:
            return urljoin(base_url, href.group(1))
    return None


def _looks_like_feed(text: str) -> bool:
    t = text.lstrip()[:500]
    return bool(re.search(r'<(rss|feed|channel)[^>]*>', t, re.IGNORECASE))


def _norm_col(name: str) -> str:
    name = re.sub(r"\s*\([^)]*\)", "", name)
    return name.strip().lower().replace(" ", "_").replace("/", "_").replace("-", "_")


async def discover_all(sources: list[dict]) -> dict[str, str]:
    """Return {source_name: feed_url} for all sources we can find feeds for."""
    results: dict[str, str] = {}

    # Apply hardcoded known feeds first
    for s in sources:
        name = s["name"]
        if name in KNOWN_FEEDS:
            results[name] = KNOWN_FEEDS[name]

    # Probe sources without a known feed
    to_probe = [
        s for s in sources
        if s["name"] not in results
        and s["name"] not in NO_FEED_SOURCES
        and s["type"] not in ("LinkedIn",)
        and s["url"]
    ]

    log.info("Probing %d sources for RSS feeds ...", len(to_probe))

    limits = httpx.Limits(max_connections=20, max_keepalive_connections=10)
    async with httpx.AsyncClient(limits=limits) as client:
        tasks = {s["name"]: probe_feed(client, s["url"]) for s in to_probe}
        done = await asyncio.gather(*tasks.values(), return_exceptions=True)
        for name, result in zip(tasks.keys(), done):
            if isinstance(result, str):
                results[name] = result
                log.info("  FOUND  %-55s -> %s", name, result)
            elif result is None:
                log.info("  no feed %-55s", name)
            else:
                log.warning("  error  %-55s: %s", name, result)

    return results


def update_excel(xlsx_path: Path, feed_map: dict[str, str], dry_run: bool = False) -> int:
    """Write feed: <url> into Notes column for each found source. Returns update count."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Sources"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hi = {}
    for i, h in enumerate(headers):
        if h:
            hi[_norm_col(str(h))] = i + 1  # 1-based column number

    name_col = hi.get("source_name")
    notes_col = hi.get("notes")
    if not name_col or not notes_col:
        raise ValueError(f"Cannot find Source Name / Notes columns. Got: {list(hi.keys())}")

    updated = 0
    for row in ws.iter_rows(min_row=2):
        name_cell = row[name_col - 1]
        notes_cell = row[notes_col - 1]
        name = str(name_cell.value or "").strip()
        existing = str(notes_cell.value or "").strip()

        if name not in feed_map:
            continue
        feed_url = feed_map[name]

        # If Notes already has a feed: hint, skip (don't overwrite manual edits)
        if "feed:" in existing.lower():
            log.info("  SKIP (already has feed:) %s", name)
            continue

        new_notes = f"feed: {feed_url}"
        if existing:
            new_notes = f"{existing}; {new_notes}"

        if not dry_run:
            notes_cell.value = new_notes
        log.info("  UPDATE %-50s -> %s", name, feed_url)
        updated += 1

    if not dry_run:
        wb.save(xlsx_path)
        log.info("Saved %s", xlsx_path)
    else:
        log.info("DRY RUN — no changes written (%d would update)", updated)

    return updated


def load_sources(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sources"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]

    def norm(h):
        if not h:
            return ""
        return re.sub(r"\s*\([^)]*\)", "", str(h)).strip().lower().replace(" ", "_").replace("/", "_").replace("-", "_")

    hi = {norm(h): i for i, h in enumerate(headers)}

    sources = []
    for r in rows[1:]:
        active = str(r[hi.get("active", 4)] or "").strip().lower()
        if active != "yes":
            continue
        sources.append({
            "name": str(r[hi.get("source_name", 0)] or "").strip(),
            "type": str(r[hi.get("type", 1)] or "").strip(),
            "url":  str(r[hi.get("url", 8)] or "").strip(),
            "notes": str(r[hi.get("notes", 24)] or "").strip(),
        })
    wb.close()
    return sources


async def main_async(xlsx_path: Path, dry_run: bool) -> None:
    sources = load_sources(xlsx_path)
    log.info("Loaded %d active sources", len(sources))

    feed_map = await discover_all(sources)
    log.info("\nDiscovered %d feeds total", len(feed_map))

    updated = update_excel(xlsx_path, feed_map, dry_run=dry_run)
    log.info("Updated %d rows in %s", updated, xlsx_path.name)

    # Print summary table
    print("\n=== FEED DISCOVERY SUMMARY ===")
    print(f"{'Source':<52} {'Feed URL'}")
    print("-" * 120)
    for name, url in sorted(feed_map.items()):
        print(f"  {name:<50}  {url}")

    no_feed = [s["name"] for s in sources
               if s["name"] not in feed_map
               and s["name"] not in NO_FEED_SOURCES
               and s["type"] not in ("LinkedIn", "ASX")]
    print(f"\n=== NO FEED FOUND ({len(no_feed)}) ===")
    for n in sorted(no_feed):
        print(f"  {n}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true",
                        help="Discover feeds but do not write to Excel")
    parser.add_argument("--xlsx", default="data/australian_defence_source_universe.xlsx")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent.parent
    xlsx = root / args.xlsx

    if not xlsx.exists():
        print(f"ERROR: {xlsx} not found")
        sys.exit(1)

    asyncio.run(main_async(xlsx, args.dry_run))
