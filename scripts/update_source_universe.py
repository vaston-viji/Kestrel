"""One-off script: append 30 source rows + 1 QA Rules row per update_defence_source_universe.md."""
import copy
import csv
import io
import sys
from pathlib import Path

import openpyxl

CSV_DATA = """Source Name,Type,Sector,Adjacent Domain,Active,Country,HQ State/Territory,Entity Level,URL,LinkedIn URL,ASX Ticker,Ownership / Affiliation,Primary or Secondary,Official Status,Relevance to Australian Defence,Topic Coverage,Update Frequency,Trust Score (1-5),Signal Score (1-5),Noise Score (1-5),Priority Tier (1-5),Include in Morning Digest,Include in Afternoon Digest,Why It Matters,Notes,Last Verified Date
Lockheed Martin Australia,Company,Defence Industry,"GWEO, guided weapons, air and missile defence, C4ISR",Yes,Australia,ACT,Company,https://www.lockheedmartin.com/en-au/index.html,,,US-parented (Lockheed Martin Corporation),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,"GWEO, guided weapons, air and missile defence, C4ISR",Weekly,5,5,2,1,Yes,Yes,"Prime contractor across GWEO, guided weapons and air/missile defence; frequently first to announce major program milestones and sovereign manufacturing investment.",ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
Boeing Defence Australia,Company,Defence Industry,"Uncrewed systems (MQ-28 Ghost Bat), airpower, sustainment",Yes,Australia,QLD,Company,https://www.boeing.com.au/,,,US-parented (The Boeing Company),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,"Uncrewed systems (MQ-28 Ghost Bat), airpower, sustainment",Weekly,5,5,2,1,Yes,Yes,Leads MQ-28 Ghost Bat development locally; major airpower sustainment prime; announcements often precede official Defence releases.,ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
BAE Systems Australia,Company,Defence Industry,"Hunter class frigate, land systems, cyber",Yes,Australia,VIC,Company,https://www.baesystems.com/en-aus/our-company/baes-australia,,,UK-parented (BAE Systems plc),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,"Hunter class frigate, land systems, cyber",Weekly,5,5,2,1,Yes,Yes,Builds the Hunter class frigate program; one of the largest single defence industrial workforces in Australia.,ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
Thales Australia,Company,Defence Industry,"Munitions (Benalla/Lithgow), sonar, protected vehicles",Yes,Australia,NSW,Company,https://www.thalesgroup.com/en/countries/asia-pacific/australia,,,French-parented (Thales Group),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,"Munitions (Benalla/Lithgow), sonar, protected vehicles",Weekly,5,5,2,1,Yes,Yes,Operates sovereign munitions and weapons manufacturing; regular first-party source on capacity expansion and GWEO-adjacent contracts.,ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
Rheinmetall Defence Australia,Company,Defence Industry,"Land 400 Boxer CRV, military vehicle manufacturing",Yes,Australia,QLD,Company,https://www.rheinmetall-defence.com.au/,,,German-parented (Rheinmetall AG),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,"Land 400 Boxer CRV, military vehicle manufacturing",Weekly,5,5,2,1,Yes,Yes,Leads Boxer CRV under Land 400 Phase 2; Military Vehicle Centre of Excellence is a flagship sovereign manufacturing case study.,ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
Saab Australia,Company,Defence Industry,"Combat management systems, naval systems",Yes,Australia,SA,Company,https://www.saab.com/markets/australia,,,Swedish-parented (Saab AB),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,"Combat management systems, naval systems",Weekly,5,5,2,1,Yes,Yes,Supplies combat management systems across the surface fleet; material to naval capability announcements.,ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
Elbit Systems of Australia,Company,Defence Industry,"Battle Management Systems, C4I, soldier systems",Yes,Australia,WA,Company,https://elbitsystems.com.au/,,,Israeli-parented (Elbit Systems Ltd),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,"Battle Management Systems, C4I, soldier systems",Weekly,5,5,2,1,Yes,Yes,Delivers Army Battle Management System; relevant to land force digitisation announcements.,ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
Leidos Australia,Company,Defence Industry,"Systems integration, sustainment, cyber",Yes,Australia,ACT,Company,https://www.leidos.com/company/international/australia,,,US-parented (Leidos Holdings),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,"Systems integration, sustainment, cyber",Weekly,5,5,2,1,Yes,Yes,Major systems integration and sustainment prime across multiple domains.,ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
Raytheon Australia,Company,Defence Industry,"Air and missile defence, guided weapons",Yes,Australia,ACT,Company,https://www.rtx.com/raytheon/who-we-are/global-locations/australia,,,US-parented (RTX Corporation),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,"Air and missile defence, guided weapons",Weekly,5,5,2,1,Yes,Yes,Delivers air and missile defence and guided weapons capability; relevant to GWEO and NASAMS-related announcements.,ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
Babcock Australasia,Company,Defence Industry,Naval and aviation sustainment,Yes,Australia,NSW,Company,https://www.babcock.com.au/,,,UK-parented (Babcock International Group),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,Naval and aviation sustainment,Weekly,5,5,2,1,Yes,Yes,Major sustainment prime across naval and aviation platforms.,ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
General Dynamics Land Systems - Australia,Company,Defence Industry,Land 400 Phase 3 Redback IFV,Yes,Australia,QLD,Company,https://www.gdlscanada.com/en/redback.html,,,US-parented (General Dynamics Corporation),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,Land 400 Phase 3 Redback IFV,Weekly,5,5,2,1,Yes,Yes,"Delivers Redback IFV under Land 400 Phase 3, a top-tier land capability program.",ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
Naval Group Australia,Company,Defence Industry,"Submarine industrial base, AUKUS transition support",Yes,Australia,SA,Company,https://www.naval-group.com/en/australia,,,French-parented (Naval Group),Primary,Official,Tier-1 global prime with a substantial Australian sovereign program; excluded under current QA Rule 2 (Australian-headquartered only) but material to contracts and market-moves reporting.,"Submarine industrial base, AUKUS transition support",Weekly,5,5,2,1,Yes,Yes,Retains a residual but relevant industrial and workforce role through the Attack class transition to AUKUS; useful for continuity and contract wind-down reporting.,ADDED AS NAMED EXCEPTION to QA Rule 2 - see QA Rules sheet update. Verify current AU newsroom/press-release URL before activating.,NEEDS VERIFICATION
Critical Minerals Facilitation Office,Government,Adjacent Domain,Critical Minerals,Yes,Australia,ACT,Commonwealth agency,https://www.industry.gov.au/science-technology-and-innovation/critical-minerals,,,Australian Government,Primary,Official,"No source in the universe currently covers critical minerals, despite it being a named category in the Email Logic template (Section 5).","Critical minerals, sovereign supply chains, processing investment",Weekly,4,4,2,2,Yes,Yes,Fills a named coverage gap (critical minerals) with zero prior representation in the source list.,VERIFY exact URL - office structure/URL may have changed.,NEEDS VERIFICATION
Defence Export Controls,Government,Defence Industry,Export Controls,Yes,Australia,ACT,Commonwealth agency,https://www.defence.gov.au/business-industry/export,,,Australian Government,Primary,Official,Export controls is a named escalation trigger in Collection Logic Step 7 but has no dedicated source.,"Defence and Strategic Goods List (DSGL), defence trade controls, export permits",Weekly,4,4,2,2,Yes,Yes,Direct source for export-control actions and DSGL changes rather than relying on incidental mentions in general Defence releases.,VERIFY exact URL and page structure.,NEEDS VERIFICATION
Incat Tasmania,Company,Defence Industry,"Maritime, high-speed vessels",Yes,Australia,TAS,Company,https://www.incat.com.au/,,,Australian,Primary,Official,Only Tasmania-headquartered entity being added; closes the one state with zero current representation.,"High-speed catamaran shipbuilding, maritime logistics, allied vessel contracts",Monthly,3,3,2,3,Yes,Yes,Fills the Tasmania gap; relevant to maritime logistics and allied vessel procurement news.,Verify current news/press page URL.,NEEDS VERIFICATION
Liferaft Systems Australia,Company,Defence Industry,"Submarine escape, maritime safety systems",Yes,Australia,TAS,Company,https://www.liferaftsystems.com/,,,Australian,Primary,Official,Niche but genuine defence-relevant Tasmanian manufacturer supplying submarine escape systems to allied navies.,"Submarine escape systems, maritime safety equipment",Monthly,3,3,2,3,Yes,Yes,Second Tasmanian entry; relevant to submarine sustainment and allied navy supply relationships.,Verify current news/press page URL.,NEEDS VERIFICATION
Team Defence Australia,Government,Defence Industry,Export Promotion,Yes,Australia,ACT,Commonwealth initiative,https://www.austrade.gov.au/en/international/team-defence-australia,,,Australian Government,Primary,Official,Named body for promoting Australian defence exports; no export-promotion source currently exists in the universe.,"Export promotion, trade missions, industry capability showcases",Monthly,4,3,2,3,Yes,Yes,Fills export-promotion gap; relevant to capture opportunities criterion in Collection Logic Step 5.,VERIFY current URL - may sit under Austrade or Defence domain.,NEEDS VERIFICATION
Export Finance Australia,Government,Adjacent Domain,Export Finance,Yes,Australia,NSW,Statutory agency,https://www.exportfinance.gov.au/,,,Australian Government,Primary,Official,Finances sovereign defence export deals; relevant to capital and contract-financing announcements.,"Export finance, sovereign defence deal financing",Monthly,4,3,2,3,Yes,Yes,Fills industrial-finance gap adjacent to Defence Industry Development Grants Program.,,NEEDS VERIFICATION
National Reconstruction Fund Corporation,Government,Adjacent Domain,Sovereign Manufacturing,Yes,Australia,ACT,Statutory agency,https://www.nrfc.gov.au/,,,Australian Government,Primary,Official,Funds sovereign manufacturing capability including defence-adjacent categories; no equivalent source currently tracked.,"Sovereign manufacturing funding, critical technologies investment",Monthly,4,3,2,3,Yes,Yes,Fills industrial-policy funding gap alongside ASCA and the Defence Industry Development Grants Program.,,NEEDS VERIFICATION
XTEK Limited,Company,Defence Industry,"Protective equipment, tactical gear",Yes,Australia,ACT,Listed company channel,https://www.xtek.net/,,XTE,Australian,Primary,Official,ASX-listed defence equipment supplier not currently tracked; relevant to the market-moves section.,"Protective equipment, tactical and ballistic gear",Monthly,3,3,2,3,Yes,Yes,Closes a specific ASX-listed coverage gap identified during the review.,Verify HQ state and current IR page URL.,NEEDS VERIFICATION
AVA Risk Group,Company,Adjacent Domain,"Fibre-optic sensing, perimeter security",Yes,Australia,SA,Listed company channel,https://www.avarisk.com/,,AVA,Australian,Primary,Official,ASX-listed sensing/security company with defence-adjacent applications not currently tracked.,"Fibre-optic sensing, perimeter and critical infrastructure security",Monthly,3,3,2,3,Yes,Yes,Closes a specific ASX-listed coverage gap identified during the review.,Verify HQ state and current IR page URL.,NEEDS VERIFICATION
Richard Marles X,X,Defence,,Yes,Australia,VIC,Minister,https://x.com/RichardMarlesMP,,,Australian,Primary,Official,Faster-breaking parallel channel to an existing LinkedIn/official source.,"Ministerial and agency statements, breaking commentary",Several times weekly,4,4,3,2,Yes,Yes,High-value source for Deputy PM & Minister for Defence statements; X is typically faster than LinkedIn for breaking remarks.,VERIFY HANDLE before activating - not live-confirmed. Mirrors existing LinkedIn entry for the same individual/agency.,NEEDS VERIFICATION
Pat Conroy X,X,Defence,,Yes,Australia,NSW,Minister,https://x.com/Pat_ConroyMP,,,Australian,Primary,Official,Faster-breaking parallel channel to an existing LinkedIn/official source.,"Ministerial and agency statements, breaking commentary",Several times weekly,4,4,3,2,Yes,Yes,High-value source for Minister for Defence Industry; parallels existing LinkedIn entry.,VERIFY HANDLE before activating - not live-confirmed. Mirrors existing LinkedIn entry for the same individual/agency.,NEEDS VERIFICATION
Matt Keogh X,X,Defence,,Yes,Australia,WA,Minister,https://x.com/MattKeoghMP,,,Australian,Primary,Official,Faster-breaking parallel channel to an existing LinkedIn/official source.,"Ministerial and agency statements, breaking commentary",Several times weekly,4,4,3,2,Yes,Yes,Parallels existing LinkedIn entry for Minister for Veterans Affairs and Defence Personnel.,VERIFY HANDLE before activating - not live-confirmed. Mirrors existing LinkedIn entry for the same individual/agency.,NEEDS VERIFICATION
Defence Australia X,X,Defence,,Yes,Australia,ACT,Agency page,https://x.com/AusDeptDefence,,,Australian,Primary,Official,Faster-breaking parallel channel to an existing LinkedIn/official source.,"Ministerial and agency statements, breaking commentary",Several times weekly,4,4,3,2,Yes,Yes,Official Defence channel; typically fastest public channel for breaking statements ahead of press release publication.,VERIFY HANDLE before activating - not live-confirmed. Mirrors existing LinkedIn entry for the same individual/agency.,NEEDS VERIFICATION
Australian Submarine Agency X,X,Defence,,Yes,Australia,ACT,Agency page,https://x.com/AusSubmarines,,,Australian,Primary,Official,Faster-breaking parallel channel to an existing LinkedIn/official source.,"Ministerial and agency statements, breaking commentary",Several times weekly,4,4,3,2,Yes,Yes,Parallels existing LinkedIn entry; AUKUS submarine program updates.,VERIFY HANDLE before activating - not live-confirmed. Mirrors existing LinkedIn entry for the same individual/agency.,NEEDS VERIFICATION
ASPI / The Strategist X,X,Defence,,Yes,Australia,ACT,Think tank publication,https://x.com/ASPI_org,,,Australian,Primary,Official,Faster-breaking parallel channel to an existing LinkedIn/official source.,"Ministerial and agency statements, breaking commentary",Several times weekly,4,4,3,2,Yes,Yes,Fast channel for ASPI commentary and report releases ahead of website publication.,VERIFY HANDLE before activating - not live-confirmed. Mirrors existing LinkedIn entry for the same individual/agency.,NEEDS VERIFICATION
The Nightly,Media,Defence,,Yes,Australia,WA,National media,https://thenightly.com.au/,,,Australian,Primary,Official,Adds channel diversity (wire service / broadcast / new masthead) without duplicating existing national media.,General defence and national security coverage,Daily,3,3,3,2,Yes,Yes,New national digital masthead (Seven West Media); explicitly requested addition.,Verify current defence-specific section/tag URL if one exists.,NEEDS VERIFICATION
AAP (Australian Associated Press),Media,Defence,,Yes,Australia,NSW,Wire service,https://aap.com.au/,,,Australian,Primary,Official,Adds channel diversity (wire service / broadcast / new masthead) without duplicating existing national media.,General defence and national security coverage,Daily,3,3,3,2,Yes,Yes,Primary Australian wire service; often the first point of publication that other outlets republish from.,Verify current defence-specific section/tag URL if one exists.,NEEDS VERIFICATION
Sky News Australia Defence,Media,Defence,,Yes,Australia,NSW,Broadcast media,https://www.skynews.com.au/,,,Australian,Primary,Official,Adds channel diversity (wire service / broadcast / new masthead) without duplicating existing national media.,General defence and national security coverage,Daily,3,3,3,2,Yes,Yes,Only broadcast/TV channel in the universe; adds a channel type not otherwise represented.,Verify current defence-specific section/tag URL if one exists.,NEEDS VERIFICATION"""

QA_CHECK = "Named exceptions to Australian-headquartered rule"
QA_HOW = (
    "Rule 2 (Australian-headquartered only) excludes offshore-parented company pages by default. "
    "Twelve Tier-1 global prime contractors Australian entities (Lockheed Martin Australia, Boeing Defence Australia, "
    "BAE Systems Australia, Thales Australia, Rheinmetall Defence Australia, Saab Australia, Elbit Systems of Australia, "
    "Leidos Australia, Raytheon Australia, Babcock Australasia, General Dynamics Land Systems - Australia, Naval Group Australia) "
    "are named exceptions given their materiality to major contract and sovereign-manufacturing announcements. "
    "Any future addition of an offshore-parented entity should be added explicitly to this exception list, "
    "not treated as a silent default."
)

root = Path(__file__).resolve().parent.parent
wb_path = root / "data" / "australian_defence_source_universe.xlsx"

wb = openpyxl.load_workbook(wb_path)
ws = wb["Sources"]

headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
print(f"Header columns found: {len(headers)}")

reader = csv.DictReader(io.StringIO(CSV_DATA))
new_rows = list(reader)
print(f"CSV rows to append: {len(new_rows)}")

template_row_idx = ws.max_row
NUMERIC_COLS = {"Trust Score (1-5)", "Signal Score (1-5)", "Noise Score (1-5)", "Priority Tier (1-5)"}


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


start_row = ws.max_row + 1
for i, row_data in enumerate(new_rows):
    row_idx = start_row + i
    for col_name, col_idx in headers.items():
        src_cell = ws.cell(template_row_idx, col_idx)
        dst_cell = ws.cell(row_idx, col_idx)
        raw = (row_data.get(col_name) or "").strip()
        if col_name in NUMERIC_COLS:
            try:
                dst_cell.value = int(raw)
            except (ValueError, TypeError):
                dst_cell.value = raw or None
        else:
            dst_cell.value = raw or None
        copy_cell_style(src_cell, dst_cell)

print(f"Sources: written rows {start_row} to {ws.max_row} ({ws.max_row - start_row + 1} rows)")

# --- QA Rules ---
qa = wb["QA Rules"]
qa_template = qa.max_row
qa_next = qa.max_row + 1
for col_idx in range(1, qa.max_column + 1):
    src = qa.cell(qa_template, col_idx)
    dst = qa.cell(qa_next, col_idx)
    copy_cell_style(src, dst)
qa.cell(qa_next, 1).value = QA_CHECK
qa.cell(qa_next, 2).value = QA_HOW
print(f"QA Rules: row {qa_next} written")

wb.save(wb_path)
print(f"Saved: {wb_path}")
